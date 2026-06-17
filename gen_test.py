import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "sheet1"
headers = ["查询类型", "姓名", "身份证号", "性别", "出生日期", "电话", "诊断", "就诊医院", "就诊科室", "就诊日期", "就诊类型"]
for col, h in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=h)
data = [
    ["medical_insurance", "张三", "430102199001011234", "男", "1990-01-01", "13800138001", "高血压", "湘雅医院", "心内科", "2025-03-15", "门诊"],
    ["medical_insurance", "李四", "440305199505052345", "女", "1995-05-05", "13900139002", "上呼吸道感染", "深圳市人民医院", "呼吸科", "2025-06-01", "门诊"],
    ["medical_record", "张三", "430102199001011234", "男", "1990-01-01", "13800138001", "高血压3级高危", "湘雅医院", "心内科", "2025-03-10", "住院"],
]
for row_idx, row_data in enumerate(data, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)
wb.save("D:\\work\\proj2\\test_import.xlsx")
print("OK")
